#!/usr/bin/env python3
"""
GhostPress Report Generator
Generates HTML and XLSX reports with detailed findings
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("Warning: openpyxl not installed - XLSX reports will not be generated")

try:
    from jinja2 import Template
    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False
    print("Warning: jinja2 not installed - using basic HTML template")


def load_findings(output_dir):
    """Load findings from JSON file"""
    findings_file = Path(output_dir) / "reports" / "findings.json"

    if not findings_file.exists():
        print(f"Error: Findings file not found: {findings_file}")
        return None

    with open(findings_file, 'r') as f:
        return json.load(f)


def get_severity_color(severity):
    """Get color code for severity level"""
    colors = {
        'CRITICAL': '#DC3545',  # Red
        'HIGH': '#FD7E14',      # Orange
        'MEDIUM': '#FFC107',    # Yellow
        'LOW': '#17A2B8',       # Blue
        'INFO': '#6C757D'       # Gray
    }
    return colors.get(severity, '#6C757D')


def get_severity_weight(severity):
    """Get numeric weight for severity (for sorting)"""
    weights = {
        'CRITICAL': 5,
        'HIGH': 4,
        'MEDIUM': 3,
        'LOW': 2,
        'INFO': 1
    }
    return weights.get(severity, 0)


def generate_html_report(output_dir, target, findings_data):
    """Generate HTML report"""
    print("Generating HTML report...")

    findings = findings_data.get('findings', [])
    metadata = findings_data.get('scan_metadata', {})

    # Sort findings by severity
    findings.sort(key=lambda x: get_severity_weight(x.get('severity', 'INFO')), reverse=True)

    # Count findings by severity
    severity_counts = {
        'CRITICAL': 0,
        'HIGH': 0,
        'MEDIUM': 0,
        'LOW': 0,
        'INFO': 0
    }

    for finding in findings:
        sev = finding.get('severity', 'INFO')
        if sev in severity_counts:
            severity_counts[sev] += 1

    # Calculate risk score
    risk_score = (
        severity_counts['CRITICAL'] * 10 +
        severity_counts['HIGH'] * 7 +
        severity_counts['MEDIUM'] * 4 +
        severity_counts['LOW'] * 2 +
        severity_counts['INFO'] * 1
    )

    if risk_score > 50:
        risk_level = "Critical"
        risk_color = "#DC3545"
    elif risk_score > 30:
        risk_level = "High"
        risk_color = "#FD7E14"
    elif risk_score > 15:
        risk_level = "Medium"
        risk_color = "#FFC107"
    else:
        risk_level = "Low"
        risk_color = "#28A745"

    # HTML template
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>GhostPress Security Assessment Report - {target}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            background: #f5f7fa;
            padding: 20px;
        }}

        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 700;
        }}

        .header .subtitle {{
            font-size: 1.2em;
            opacity: 0.9;
        }}

        .metadata {{
            background: #f8f9fa;
            padding: 20px 40px;
            border-bottom: 1px solid #dee2e6;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
        }}

        .metadata-item {{
            display: flex;
            flex-direction: column;
        }}

        .metadata-label {{
            font-size: 0.85em;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 5px;
        }}

        .metadata-value {{
            font-size: 1.1em;
            font-weight: 600;
            color: #495057;
        }}

        .content {{
            padding: 40px;
        }}

        .section {{
            margin-bottom: 40px;
        }}

        .section-title {{
            font-size: 1.8em;
            margin-bottom: 20px;
            color: #2c3e50;
            border-bottom: 3px solid #667eea;
            padding-bottom: 10px;
        }}

        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}

        .stat-card {{
            background: white;
            border: 2px solid #e9ecef;
            border-radius: 8px;
            padding: 20px;
            text-align: center;
            transition: transform 0.2s, box-shadow 0.2s;
        }}

        .stat-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }}

        .stat-card.critical {{ border-color: #DC3545; }}
        .stat-card.high {{ border-color: #FD7E14; }}
        .stat-card.medium {{ border-color: #FFC107; }}
        .stat-card.low {{ border-color: #17A2B8; }}
        .stat-card.info {{ border-color: #6C757D; }}

        .stat-number {{
            font-size: 3em;
            font-weight: 700;
            margin-bottom: 5px;
        }}

        .stat-number.critical {{ color: #DC3545; }}
        .stat-number.high {{ color: #FD7E14; }}
        .stat-number.medium {{ color: #FFC107; }}
        .stat-number.low {{ color: #17A2B8; }}
        .stat-number.info {{ color: #6C757D; }}

        .stat-label {{
            font-size: 0.9em;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}

        .risk-assessment {{
            background: linear-gradient(135deg, {risk_color}15 0%, {risk_color}05 100%);
            border-left: 5px solid {risk_color};
            padding: 25px;
            border-radius: 8px;
            margin-bottom: 30px;
        }}

        .risk-assessment h3 {{
            color: {risk_color};
            margin-bottom: 10px;
            font-size: 1.5em;
        }}

        .risk-level {{
            font-size: 2em;
            font-weight: 700;
            color: {risk_color};
        }}

        .finding {{
            background: white;
            border: 1px solid #e9ecef;
            border-radius: 8px;
            padding: 25px;
            margin-bottom: 20px;
            transition: box-shadow 0.2s;
        }}

        .finding:hover {{
            box-shadow: 0 4px 15px rgba(0,0,0,0.08);
        }}

        .finding-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 15px;
            flex-wrap: wrap;
            gap: 10px;
        }}

        .finding-title {{
            font-size: 1.4em;
            font-weight: 600;
            color: #2c3e50;
            flex: 1;
        }}

        .severity-badge {{
            display: inline-block;
            padding: 5px 15px;
            border-radius: 20px;
            font-size: 0.85em;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            color: white;
        }}

        .severity-badge.critical {{ background: #DC3545; }}
        .severity-badge.high {{ background: #FD7E14; }}
        .severity-badge.medium {{ background: #FFC107; color: #333; }}
        .severity-badge.low {{ background: #17A2B8; }}
        .severity-badge.info {{ background: #6C757D; }}

        .category-badge {{
            display: inline-block;
            padding: 5px 12px;
            border-radius: 5px;
            font-size: 0.8em;
            background: #e9ecef;
            color: #495057;
            margin-left: 10px;
        }}

        .finding-section {{
            margin-bottom: 15px;
        }}

        .finding-section-title {{
            font-weight: 700;
            color: #495057;
            margin-bottom: 8px;
            font-size: 1.05em;
            display: flex;
            align-items: center;
        }}

        .finding-section-title::before {{
            content: '';
            width: 4px;
            height: 18px;
            background: #667eea;
            margin-right: 10px;
            border-radius: 2px;
        }}

        .finding-section-content {{
            color: #666;
            line-height: 1.7;
            padding-left: 14px;
        }}

        .evidence {{
            background: #f8f9fa;
            border-left: 3px solid #667eea;
            padding: 15px;
            border-radius: 5px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
            overflow-x: auto;
            color: #495057;
        }}

        .footer {{
            background: #f8f9fa;
            padding: 30px 40px;
            border-top: 1px solid #dee2e6;
            text-align: center;
            color: #6c757d;
        }}

        .footer a {{
            color: #667eea;
            text-decoration: none;
        }}

        @media print {{
            body {{
                background: white;
                padding: 0;
            }}

            .container {{
                box-shadow: none;
            }}

            .finding {{
                page-break-inside: avoid;
            }}
        }}

        @media (max-width: 768px) {{
            .header {{
                padding: 30px 20px;
            }}

            .header h1 {{
                font-size: 2em;
            }}

            .content {{
                padding: 20px;
            }}

            .stats-grid {{
                grid-template-columns: repeat(2, 1fr);
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🔒 GhostPress Security Assessment</h1>
            <div class="subtitle">WordPress Vulnerability Report</div>
        </div>

        <div class="metadata">
            <div class="metadata-item">
                <div class="metadata-label">Target</div>
                <div class="metadata-value">{target}</div>
            </div>
            <div class="metadata-item">
                <div class="metadata-label">Scan Date</div>
                <div class="metadata-value">{metadata.get('scan_date', 'N/A')}</div>
            </div>
            <div class="metadata-item">
                <div class="metadata-label">Tool Version</div>
                <div class="metadata-value">GhostPress v{metadata.get('tool_version', '2.0')}</div>
            </div>
            <div class="metadata-item">
                <div class="metadata-label">Total Findings</div>
                <div class="metadata-value">{len(findings)}</div>
            </div>
        </div>

        <div class="content">
            <div class="section">
                <h2 class="section-title">Executive Summary</h2>

                <div class="risk-assessment">
                    <h3>Overall Risk Assessment</h3>
                    <div class="risk-level">{risk_level}</div>
                    <p>Risk Score: {risk_score} / 100</p>
                </div>

                <div class="stats-grid">
                    <div class="stat-card critical">
                        <div class="stat-number critical">{severity_counts['CRITICAL']}</div>
                        <div class="stat-label">Critical</div>
                    </div>
                    <div class="stat-card high">
                        <div class="stat-number high">{severity_counts['HIGH']}</div>
                        <div class="stat-label">High</div>
                    </div>
                    <div class="stat-card medium">
                        <div class="stat-number medium">{severity_counts['MEDIUM']}</div>
                        <div class="stat-label">Medium</div>
                    </div>
                    <div class="stat-card low">
                        <div class="stat-number low">{severity_counts['LOW']}</div>
                        <div class="stat-label">Low</div>
                    </div>
                    <div class="stat-card info">
                        <div class="stat-number info">{severity_counts['INFO']}</div>
                        <div class="stat-label">Info</div>
                    </div>
                </div>
            </div>
"""

    # Add findings by severity
    for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO']:
        severity_findings = [f for f in findings if f.get('severity') == severity]

        if severity_findings:
            html_content += f"""
            <div class="section">
                <h2 class="section-title">{severity} Severity Findings ({len(severity_findings)})</h2>
"""

            for finding in severity_findings:
                title = finding.get('title', 'Untitled Finding')
                description = finding.get('description', 'No description provided')
                impact = finding.get('impact', 'Impact not specified')
                mitigation = finding.get('mitigation', 'Mitigation not specified')
                category = finding.get('category', 'General')
                evidence = finding.get('evidence', 'N/A')

                html_content += f"""
                <div class="finding">
                    <div class="finding-header">
                        <div class="finding-title">{title}</div>
                        <div>
                            <span class="severity-badge {severity.lower()}">{severity}</span>
                            <span class="category-badge">{category}</span>
                        </div>
                    </div>

                    <div class="finding-section">
                        <div class="finding-section-title">Description</div>
                        <div class="finding-section-content">{description}</div>
                    </div>

                    <div class="finding-section">
                        <div class="finding-section-title">Impact</div>
                        <div class="finding-section-content">{impact}</div>
                    </div>

                    <div class="finding-section">
                        <div class="finding-section-title">Remediation</div>
                        <div class="finding-section-content">{mitigation}</div>
                    </div>

                    <div class="finding-section">
                        <div class="finding-section-title">Evidence</div>
                        <div class="evidence">{evidence}</div>
                    </div>
                </div>
"""

            html_content += """
            </div>
"""

    html_content += f"""
        </div>

        <div class="footer">
            <p><strong>GhostPress Security Assessment Tool</strong></p>
            <p>Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>This report is confidential and intended for authorized use only.</p>
        </div>
    </div>
</body>
</html>
"""

    # Write HTML file
    html_file = Path(output_dir) / "reports" / "report.html"
    with open(html_file, 'w') as f:
        f.write(html_content)

    print(f"✓ HTML report generated: {html_file}")


def generate_xlsx_report(output_dir, target, findings_data):
    """Generate Excel (XLSX) report"""
    if not XLSX_AVAILABLE:
        print("✗ XLSX report generation skipped (openpyxl not installed)")
        return

    print("Generating XLSX report...")

    findings = findings_data.get('findings', [])
    metadata = findings_data.get('scan_metadata', {})

    # Sort findings by severity
    findings.sort(key=lambda x: get_severity_weight(x.get('severity', 'INFO')), reverse=True)

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary")

    # Header styling
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Title
    ws_summary.merge_cells('A1:D1')
    title_cell = ws_summary['A1']
    title_cell.value = 'GhostPress Security Assessment Report'
    title_cell.font = Font(name='Calibri', size=18, bold=True)
    title_cell.alignment = header_alignment
    ws_summary.row_dimensions[1].height = 30

    # Metadata
    ws_summary['A3'] = 'Target:'
    ws_summary['B3'] = target
    ws_summary['A4'] = 'Scan Date:'
    ws_summary['B4'] = metadata.get('scan_date', 'N/A')
    ws_summary['A5'] = 'Tool Version:'
    ws_summary['B5'] = f"GhostPress v{metadata.get('tool_version', '2.0')}"
    ws_summary['A6'] = 'Total Findings:'
    ws_summary['B6'] = len(findings)

    # Bold labels
    for row in range(3, 7):
        ws_summary[f'A{row}'].font = Font(bold=True)

    # Severity counts
    ws_summary['A8'] = 'Findings by Severity'
    ws_summary['A8'].font = Font(size=14, bold=True)

    severity_counts = {
        'CRITICAL': 0,
        'HIGH': 0,
        'MEDIUM': 0,
        'LOW': 0,
        'INFO': 0
    }

    for finding in findings:
        sev = finding.get('severity', 'INFO')
        if sev in severity_counts:
            severity_counts[sev] += 1

    row = 9
    for severity, count in severity_counts.items():
        ws_summary[f'A{row}'] = severity
        ws_summary[f'B{row}'] = count

        # Color code
        color = get_severity_color(severity).replace('#', '')
        ws_summary[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        if severity in ['CRITICAL', 'HIGH', 'LOW', 'INFO']:
            ws_summary[f'A{row}'].font = Font(color='FFFFFF', bold=True)
        else:
            ws_summary[f'A{row}'].font = Font(bold=True)

        row += 1

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40

    # Create Findings sheet
    ws_findings = wb.create_sheet("Findings")

    # Headers
    headers = ['ID', 'Title', 'Severity', 'Category', 'Description', 'Impact', 'Mitigation', 'Evidence']
    for col, header in enumerate(headers, start=1):
        cell = ws_findings.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add findings
    for idx, finding in enumerate(findings, start=2):
        ws_findings.cell(row=idx, column=1, value=finding.get('id', ''))
        ws_findings.cell(row=idx, column=2, value=finding.get('title', ''))
        ws_findings.cell(row=idx, column=3, value=finding.get('severity', ''))
        ws_findings.cell(row=idx, column=4, value=finding.get('category', ''))
        ws_findings.cell(row=idx, column=5, value=finding.get('description', ''))
        ws_findings.cell(row=idx, column=6, value=finding.get('impact', ''))
        ws_findings.cell(row=idx, column=7, value=finding.get('mitigation', ''))
        ws_findings.cell(row=idx, column=8, value=finding.get('evidence', ''))

        # Color code severity
        severity = finding.get('severity', 'INFO')
        color = get_severity_color(severity).replace('#', '')
        severity_cell = ws_findings.cell(row=idx, column=3)
        severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        if severity in ['CRITICAL', 'HIGH', 'LOW', 'INFO']:
            severity_cell.font = Font(color='FFFFFF', bold=True)
        else:
            severity_cell.font = Font(bold=True)

        # Wrap text for long fields
        for col in [2, 5, 6, 7, 8]:
            ws_findings.cell(row=idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    # Adjust column widths
    column_widths = {
        'A': 15,  # ID
        'B': 35,  # Title
        'C': 12,  # Severity
        'D': 18,  # Category
        'E': 50,  # Description
        'F': 50,  # Impact
        'G': 50,  # Mitigation
        'H': 40   # Evidence
    }

    for col, width in column_widths.items():
        ws_findings.column_dimensions[col].width = width

    # Freeze header row
    ws_findings.freeze_panes = 'A2'

    # Add autofilter
    ws_findings.auto_filter.ref = ws_findings.dimensions

    # Create sheets by severity
    for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO']:
        severity_findings = [f for f in findings if f.get('severity') == severity]

        if severity_findings:
            ws = wb.create_sheet(f"{severity} Findings")

            # Headers
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Add findings
            for idx, finding in enumerate(severity_findings, start=2):
                ws.cell(row=idx, column=1, value=finding.get('id', ''))
                ws.cell(row=idx, column=2, value=finding.get('title', ''))
                ws.cell(row=idx, column=3, value=finding.get('severity', ''))
                ws.cell(row=idx, column=4, value=finding.get('category', ''))
                ws.cell(row=idx, column=5, value=finding.get('description', ''))
                ws.cell(row=idx, column=6, value=finding.get('impact', ''))
                ws.cell(row=idx, column=7, value=finding.get('mitigation', ''))
                ws.cell(row=idx, column=8, value=finding.get('evidence', ''))

                # Wrap text
                for col in [2, 5, 6, 7, 8]:
                    ws.cell(row=idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            # Adjust column widths
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Freeze header
            ws.freeze_panes = 'A2'

            # Add autofilter
            ws.auto_filter.ref = ws.dimensions

    # Save workbook
    xlsx_file = Path(output_dir) / "reports" / "report.xlsx"
    wb.save(xlsx_file)

    print(f"✓ XLSX report generated: {xlsx_file}")


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 generate_reports.py <output_dir> <target>")
        sys.exit(1)

    output_dir = sys.argv[1]
    target = sys.argv[2]

    print(f"\nGenerating reports for {target}...")
    print(f"Output directory: {output_dir}\n")

    # Load findings
    findings_data = load_findings(output_dir)
    if not findings_data:
        sys.exit(1)

    # Generate reports
    try:
        generate_html_report(output_dir, target, findings_data)
        generate_xlsx_report(output_dir, target, findings_data)
        print("\n✓ Report generation complete!\n")
    except Exception as e:
        print(f"\n✗ Error generating reports: {e}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
